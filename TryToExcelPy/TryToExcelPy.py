from openpyxl import*
from openpyxl.styles import*






work_book = open("DowletSanaw1_ay (1).xlsx")
work_sheet = work_book.active


new_wb = Workbook()
new_ws = new_wb.create_sheet("MySheet", 0)
new_ws["A1"] = "Город"
new_ws["B1"] = "Область/Этрап"
new_ws["C1"] = "Район"

i = 3
for row in range(3, work_sheet.max_row+1):
    value = work_sheet[f"R{i}"].value
    result = value.split()
    
    new_ws[f"A{i}"] = result[0]
    if len(result) == 1:
        new_ws[f"B{i}"] = '-'
    else:
        new_ws[f"B{i}"] = result[2]
    

   
    new_value = result[4:]
    str_new_value = ' '.join(new_value)
    new_ws[f"C{i}"] = str_new_value
    new_ws.column_dimensions['C'].width = len(str_new_value)+30
    


    i+=1



def __format_ws__(ws, cell_range):

        #applying border and alignment
        font = Font(size=9)
        align=Alignment(horizontal='left', vertical='center')
        border = Border(left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000'))

        rows = [rows for rows in ws[cell_range]]
        flattened = [item for sublist in rows for item in sublist]
        [(setattr(cell,'border',border), setattr(cell,'font',font),
         setattr(cell,'alignment',align)) for cell in flattened]


__format_ws__(ws = new_ws, cell_range = f"A1:C{work_sheet.max_row}")

new_wb.save(filename = 'NewDoc.xlsx')